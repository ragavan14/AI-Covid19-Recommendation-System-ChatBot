from covid import Covid
import pandas as pd
import openpyxl
# import os
import datetime
import matplotlib.pyplot as plt
from openpyxl import Workbook
import tkinter as tk
from tkinter import simpledialog


def graph1():
    file = Workbook()

    sheet = file.active
    sheet["A1"] = "Date"
    sheet["B1"] = "Country"
    sheet["C1"] = "Confirmed_Cases"
    sheet["D1"] = "Active_Cases"
    sheet["E1"] = "Recovered_Cases"
    sheet["F1"] = "Death_Cases"
    file.save("covid-19.xlsx")

    covid = Covid(source="worldometers")
    data = covid.get_data()

    confirmed = covid.get_total_confirmed_cases()

    active = covid.get_total_active_cases()

    recovered = covid.get_total_recovered()

    deaths = covid.get_total_deaths()

    file = openpyxl.load_workbook("covid-19.xlsx")
    sheet = file.active

    time = datetime.datetime.now().strftime("%d-%m-%y #H:#M:#S")

    for i in data:
        sheet.cell(column=1, row=sheet.max_row + 1, value=time)
        sheet.cell(column=2, row=sheet.max_row, value=i["country"])
        sheet.cell(column=3, row=sheet.max_row, value=i["confirmed"])
        sheet.cell(column=4, row=sheet.max_row, value=i["active"])
        sheet.cell(column=5, row=sheet.max_row, value=i["recovered"])
        sheet.cell(column=6, row=sheet.max_row, value=i["deaths"])

    file.save("covid-19.xlsx")

    app = tk.Tk()

    app.state("zoomed")
    app.withdraw()

    USER_INP = simpledialog.askstring(title="CHITTI", prompt="Enter the country name:")

    country = USER_INP

    for i in data:
        if i["country"] == country:
            print("country: ", i["confirmed"], "\nActive :", i["active"],
                  "\nRecovered :", i["recovered"], "\nDeaths :", i["deaths"])

    x1 = pd.read_excel('covid-19.xlsx', 'Sheet')
    x1.to_csv('Manipulation.csv', index=False)

    df = pd.read_csv('Manipulation.csv')

    pos = [0, 1, 2, 3, 4, 5, 6, 7]
    df.drop(df.index[pos], inplace=True)

    x = df.nlargest(70, 'Confirmed_Cases')

    x.plot(kind='bar', x='Country', y='Confirmed_Cases', title='Top Countries with +ve cases')
    plt.xlabel('Country')
    plt.xlabel('Confirmed cases')
    plt.show()
